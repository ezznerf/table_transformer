import re
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string


class ExcelHelper:
    @staticmethod
    def create_df(cells_dict):
        cell_data = [[None] for _ in cells_dict.keys()]
        df = pd.DataFrame(cell_data, index=list(cells_dict.keys()), columns=["Cell"])
        return df

    @staticmethod
    def create_empty_excel_file(cells_dict, file_name='empty_table_with_borders.xlsx'):
        max_col = max([column_index_from_string(ExcelHelper.split_cell_name(cell)[0]) for cell in cells_dict.keys()])
        max_row = max([ExcelHelper.split_cell_name(cell)[1] for cell in cells_dict.keys()])
        print("Макс. номер колонки:", max_col)
        print("Макс. номер строки:", max_row)
        wb = openpyxl.Workbook()
        ws = wb.active
        border_style = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border_style
        wb.save(file_name)
        print(f"Excel-файл сохранён как '{file_name}'")
        return file_name

    @staticmethod
    def split_cell_name(cell):
        match = re.match(r"([A-Z]+)(\d+)", cell)
        if match:
            col, row = match.groups()
            return col, int(row)
        else:
            raise ValueError(f"Некорректное имя ячейки: {cell}")

    @staticmethod
    def create_excel(excel_name, text_to_cells):
        file_name = 'empty_table_with_borders.xlsx'
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active

        font = Font(name='Times New Roman', size=14)
        min_column_width = 1

        for text, cells in text_to_cells.items():
            print("Обрабатываем ячейки:", cells)
            col_start = column_index_from_string(ExcelHelper.split_cell_name(cells[0])[0])
            row_start = ExcelHelper.split_cell_name(cells[0])[1]
            col_end = column_index_from_string(ExcelHelper.split_cell_name(cells[-1])[0])
            row_end = ExcelHelper.split_cell_name(cells[-1])[1]

            ws.merge_cells(start_row=row_start, start_column=col_start,
                           end_row=row_end, end_column=col_end)
            cell = ws.cell(row=row_start, column=col_start)
            if not isinstance(cell, openpyxl.cell.MergedCell):
                cell.value = text
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.font = font

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell_value = str(ws.cell(row=row, column=col_idx).value or "")
                max_length = max(max_length, len(cell_value.replace('\n', '')))
            ws.column_dimensions[col_letter].width = max(min_column_width, 10)

        for row_idx in range(1, ws.max_row + 1):
            max_text_height = 1
            for col_idx in range(1, ws.max_column + 1):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                if cell_value:
                    lines = cell_value.split('\n')
                    max_text_height = max(max_text_height, len(lines))
            ws.row_dimensions[row_idx].height = min(max_text_height * 12, 1000)

        wb.save(excel_name)
        print(f"Обработка закончена, файл '{excel_name}' создан.")
